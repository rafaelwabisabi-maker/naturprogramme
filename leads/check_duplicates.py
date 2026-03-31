#!/usr/bin/env python3
"""
Duplicate Contact Checker for Naturpädagogik Excel Files
Compares 5 Excel files for duplicate contacts (orgs and emails)
Uses fuzzy matching for organization names
"""

from openpyxl import load_workbook
from difflib import SequenceMatcher
import os
from pathlib import Path

# Configuration
BASE_PATH = "/Users/apple/Desktop/PROJECTS/school project"
FILES = [
    "Naturpaedagogik_Prospect_Database_OOe.xlsx",
    "Naturpaedagogik_Schullandwochen_OOe.xlsx",
    "Naturpaedagogik_Ferienpass_Gemeinden_OOe.xlsx",
    "Naturpaedagogik_Grenzregionen.xlsx",
    "Naturpaedagogik_Extended_1h20_1h50.xlsx",
]

# Map file index to expected org/email columns
COLUMN_MAP = {
    0: {"org": "B", "email": "G"},  # Prospect Database
    1: {"org": "B", "email": "G"},  # Schullandwochen
    2: {"org": "B", "email": "E"},  # Ferienpass (different order)
    3: {"org": "B", "email": "H"},  # Grenzregionen
    4: {"org": "B", "email": "H"},  # Extended
}

FUZZY_THRESHOLD = 0.85  # Similarity threshold for org name matching (0-1)


def col_letter_to_num(col_letter):
    """Convert column letter to number (A=1, B=2, etc.)"""
    return ord(col_letter.upper()) - ord('A')


def get_worksheet(file_path):
    """Load workbook and return first sheet"""
    wb = load_workbook(file_path)
    return wb.active


def extract_contacts(file_index, worksheet):
    """
    Extract organization and email pairs from a worksheet.
    Returns list of tuples: (org, email, row_num)
    """
    contacts = []
    col_map = COLUMN_MAP[file_index]
    org_col = col_letter_to_num(col_map["org"]) + 1
    email_col = col_letter_to_num(col_map["email"]) + 1
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) >= max(org_col, email_col):
            org = row[org_col - 1]
            email = row[email_col - 1]
            
            # Clean up data
            org = str(org).strip() if org else None
            email = str(email).strip() if email else None
            
            # Only add if we have at least organization or email
            if org and org.lower() != "none":
                contacts.append({
                    "org": org,
                    "email": email if email and "@" in email else None,
                    "row": row_idx
                })
    
    return contacts


def fuzzy_match(s1, s2, threshold=FUZZY_THRESHOLD):
    """Check if two strings are similar enough (fuzzy match)"""
    if not s1 or not s2:
        return False
    s1_lower = s1.lower().strip()
    s2_lower = s2.lower().strip()
    
    # Exact match (case-insensitive)
    if s1_lower == s2_lower:
        return True
    
    # Fuzzy match
    ratio = SequenceMatcher(None, s1_lower, s2_lower).ratio()
    return ratio >= threshold


def email_match(e1, e2):
    """Check if two emails match (case-insensitive)"""
    if not e1 or not e2:
        return False
    return e1.lower().strip() == e2.lower().strip()


def find_duplicates():
    """Main function to find duplicate contacts across all files"""
    
    print("=" * 80)
    print("DUPLICATE CONTACT CHECKER - Naturpädagogik Excel Files")
    print("=" * 80)
    print()
    
    # Load all files
    all_data = []
    total_contacts = 0
    
    for idx, filename in enumerate(FILES):
        file_path = os.path.join(BASE_PATH, filename)
        print(f"Loading {idx + 1}. {filename}...", end=" ")
        
        try:
            ws = get_worksheet(file_path)
            contacts = extract_contacts(idx, ws)
            all_data.append({
                "file_index": idx,
                "filename": filename,
                "contacts": contacts,
                "count": len(contacts)
            })
            print(f"✓ {len(contacts)} contacts found")
            total_contacts += len(contacts)
        except Exception as e:
            print(f"✗ ERROR: {e}")
            return
    
    print()
    print(f"TOTAL CONTACTS LOADED: {total_contacts}")
    print()
    
    # Check for duplicates across all files
    print("=" * 80)
    print("CROSS-FILE DUPLICATE ANALYSIS")
    print("=" * 80)
    print()
    
    duplicates_found = False
    
    # Compare each file against all others
    for i, data_i in enumerate(all_data):
        for j, data_j in enumerate(all_data):
            if i >= j:  # Skip self-comparison and reversed pairs
                continue
            
            file_i = data_i["filename"]
            file_j = data_j["filename"]
            contacts_i = data_i["contacts"]
            contacts_j = data_j["contacts"]
            
            matches = []
            
            # Check for organization name matches (fuzzy)
            for contact_i in contacts_i:
                org_i = contact_i["org"]
                for contact_j in contacts_j:
                    org_j = contact_j["org"]
                    if fuzzy_match(org_i, org_j):
                        matches.append({
                            "type": "ORG_MATCH",
                            "org_i": org_i,
                            "org_j": org_j,
                            "file_i": file_i,
                            "file_j": file_j,
                            "row_i": contact_i["row"],
                            "row_j": contact_j["row"],
                        })
            
            # Check for email matches (exact)
            for contact_i in contacts_i:
                email_i = contact_i["email"]
                if email_i:
                    for contact_j in contacts_j:
                        email_j = contact_j["email"]
                        if email_match(email_i, email_j):
                            matches.append({
                                "type": "EMAIL_MATCH",
                                "email": email_i,
                                "org_i": contact_i["org"],
                                "org_j": contact_j["org"],
                                "file_i": file_i,
                                "file_j": file_j,
                                "row_i": contact_i["row"],
                                "row_j": contact_j["row"],
                            })
            
            # Report matches
            if matches:
                duplicates_found = True
                print(f"DUPLICATES FOUND: {file_i} ↔ {file_j}")
                print("-" * 80)
                for match in matches:
                    if match["type"] == "ORG_MATCH":
                        print(f"  ORG MATCH (Row {match['row_i']} ↔ Row {match['row_j']})")
                        print(f"    {match['file_i']}: {match['org_i']}")
                        print(f"    {match['file_j']}: {match['org_j']}")
                    else:  # EMAIL_MATCH
                        print(f"  EMAIL MATCH (Row {match['row_i']} ↔ Row {match['row_j']})")
                        print(f"    Email: {match['email']}")
                        print(f"    {match['file_i']}: {match['org_i']}")
                        print(f"    {match['file_j']}: {match['org_j']}")
                    print()
                print()
    
    if not duplicates_found:
        print("✓ NO DUPLICATES FOUND across files!")
        print()
    
    # Check file 5 for internal duplicates
    print("=" * 80)
    print("FILE 5 INTERNAL DUPLICATE CHECK")
    print("=" * 80)
    print()
    
    file_5_data = all_data[4]
    file_5_contacts = file_5_data["contacts"]
    
    internal_duplicates = []
    
    # Check for duplicate emails within file 5
    emails_seen = {}
    for contact in file_5_contacts:
        if contact["email"]:
            email_lower = contact["email"].lower()
            if email_lower in emails_seen:
                internal_duplicates.append({
                    "type": "EMAIL",
                    "email": contact["email"],
                    "orgs": [emails_seen[email_lower]["org"], contact["org"]],
                    "rows": [emails_seen[email_lower]["row"], contact["row"]],
                })
            else:
                emails_seen[email_lower] = contact
    
    # Check for duplicate orgs within file 5 (fuzzy match)
    orgs_checked = set()
    for i, contact_i in enumerate(file_5_contacts):
        org_i = contact_i["org"]
        org_i_lower = org_i.lower()
        
        if org_i_lower in orgs_checked:
            continue
        orgs_checked.add(org_i_lower)
        
        for contact_j in file_5_contacts[i + 1:]:
            org_j = contact_j["org"]
            if fuzzy_match(org_i, org_j):
                internal_duplicates.append({
                    "type": "ORG",
                    "org_i": org_i,
                    "org_j": org_j,
                    "rows": [contact_i["row"], contact_j["row"]],
                })
    
    if internal_duplicates:
        print(f"⚠ INTERNAL DUPLICATES FOUND in {file_5_data['filename']}:")
        print("-" * 80)
        for dup in internal_duplicates:
            if dup["type"] == "EMAIL":
                print(f"  EMAIL DUPLICATE (Rows {dup['rows'][0]} ↔ {dup['rows'][1]})")
                print(f"    Email: {dup['email']}")
                print(f"    Org 1: {dup['orgs'][0]}")
                print(f"    Org 2: {dup['orgs'][1]}")
            else:  # ORG
                print(f"  ORG DUPLICATE (Rows {dup['rows'][0]} ↔ {dup['rows'][1]})")
                print(f"    {dup['org_i']}")
                print(f"    {dup['org_j']}")
            print()
    else:
        print("✓ NO INTERNAL DUPLICATES FOUND in File 5!")
        print()
    
    # Summary
    print("=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Files analyzed: {len(FILES)}")
    print(f"Total contacts: {total_contacts}")
    print(f"Cross-file duplicates: {'Found' if duplicates_found else 'None'}")
    print(f"File 5 internal duplicates: {len(internal_duplicates) if internal_duplicates else 'None'}")
    print()


if __name__ == "__main__":
    find_duplicates()
